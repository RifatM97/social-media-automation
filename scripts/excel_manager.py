"""
Excel Manager - handles all read/write operations for the content tracking spreadsheet.

Workbook structure:
  Sheet "Videos"  - YouTube videos fetched from monitored channels
  Sheet "Content" - Generated social media content awaiting approval
  Sheet "Config"  - Runtime settings (publish rate, platforms, last check time, etc.)
"""

import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

_SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
_PLUGIN_ROOT = os.path.dirname(_SCRIPTS_DIR)
EXCEL_PATH = os.path.join(_PLUGIN_ROOT, "data", "content_tracker.xlsx")

# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

VIDEO_HEADERS = [
    "Video ID",       # A
    "Channel Name",   # B
    "Channel ID",     # C
    "Title",          # D
    "Published Date", # E
    "URL",            # F
    "Summary",        # G
    "Status",         # H  Pending Review | Selected | Ignored | Processing | Processed | Error
    "Date Added",     # I
]

CONTENT_HEADERS = [
    "Content ID",     # A
    "Video ID",       # B
    "Video Title",    # C
    "Content Type",   # D  Tweet Thread | LinkedIn Post | Newsletter | Instagram Caption
    "Platform",       # E  Twitter/X | LinkedIn | Newsletter | Instagram
    "Content Text",   # F
    "Status",         # G  Pending Approval | Approved | Scheduled | Published | Ready for CRM | Rejected | Error
    "Scheduled Date", # H
    "Published Date", # I
    "Date Generated", # J
    "Notes",          # K
]

CONFIG_DEFAULTS = [
    ("publish_rate_per_day", "1",                              "Number of posts to publish per day"),
    ("default_platforms",    "twitter,linkedin",               "Comma-separated platforms for auto-scheduling"),
    ("content_types",        "tweet_thread,linkedin_post,newsletter", "Comma-separated content types to generate"),
    ("last_check",           "",                               "ISO timestamp of last YouTube channel check"),
]

VIDEO_COL_WIDTHS   = [18, 22, 26, 55, 22, 52, 80, 22, 22]
CONTENT_COL_WIDTHS = [30, 18, 45, 22, 18, 100, 25, 22, 22, 22, 35]
CONFIG_COL_WIDTHS  = [32, 45, 65]

HEADER_BG = "2E4057"
HEADER_FG = "FFFFFF"


# ---------------------------------------------------------------------------
# Workbook helpers
# ---------------------------------------------------------------------------

def get_workbook() -> Workbook:
    """Load existing workbook or create a fresh one."""
    if os.path.exists(EXCEL_PATH):
        return load_workbook(EXCEL_PATH)
    return _create_workbook()


def save_workbook(wb: Workbook) -> None:
    wb.save(EXCEL_PATH)


def _style_header(ws, num_cols: int) -> None:
    fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    font = Font(color=HEADER_FG, bold=True)
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)


def _set_col_widths(ws, widths: list) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _create_workbook() -> Workbook:
    wb = Workbook()

    # --- Videos sheet ---
    ws_v = wb.active
    ws_v.title = "Videos"
    ws_v.row_dimensions[1].height = 20
    ws_v.append(VIDEO_HEADERS)
    _style_header(ws_v, len(VIDEO_HEADERS))
    _set_col_widths(ws_v, VIDEO_COL_WIDTHS)

    # --- Content sheet ---
    ws_c = wb.create_sheet("Content")
    ws_c.row_dimensions[1].height = 20
    ws_c.append(CONTENT_HEADERS)
    _style_header(ws_c, len(CONTENT_HEADERS))
    _set_col_widths(ws_c, CONTENT_COL_WIDTHS)

    # --- Config sheet ---
    ws_cfg = wb.create_sheet("Config")
    ws_cfg.append(["Setting", "Value", "Description"])
    _style_header(ws_cfg, 3)
    _set_col_widths(ws_cfg, CONFIG_COL_WIDTHS)
    for row in CONFIG_DEFAULTS:
        ws_cfg.append(list(row))

    os.makedirs(os.path.dirname(EXCEL_PATH), exist_ok=True)
    wb.save(EXCEL_PATH)
    print(f"[excel_manager] Created new tracker at: {EXCEL_PATH}")
    return wb


# ---------------------------------------------------------------------------
# Videos sheet operations
# ---------------------------------------------------------------------------

def add_video(video_id: str, channel_name: str, channel_id: str,
              title: str, published_date: str, url: str, summary: str) -> bool:
    """
    Add a video row. Returns True if inserted, False if already present.
    """
    wb = get_workbook()
    ws = wb["Videos"]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == video_id:
            return False  # duplicate

    ws.append([
        video_id, channel_name, channel_id, title,
        published_date, url, summary, "Pending Review",
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    ])
    save_workbook(wb)
    return True


def get_selected_videos() -> list[dict]:
    """Return all rows with Status == 'Selected'."""
    wb = get_workbook()
    ws = wb["Videos"]
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[7] == "Selected":
            result.append({
                "video_id":       row[0],
                "channel_name":   row[1],
                "channel_id":     row[2],
                "title":          row[3],
                "published_date": row[4],
                "url":            row[5],
                "summary":        row[6],
            })
    return result


def update_video_status(video_id: str, new_status: str) -> bool:
    """Update the Status cell for a video row."""
    wb = get_workbook()
    ws = wb["Videos"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == video_id:
            row[7].value = new_status
            save_workbook(wb)
            return True
    return False


def list_videos(status_filter: str | None = None) -> list[dict]:
    """Return video rows, optionally filtered by status."""
    wb = get_workbook()
    ws = wb["Videos"]
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        if status_filter and row[7] != status_filter:
            continue
        result.append(dict(zip(VIDEO_HEADERS, row)))
    return result


# ---------------------------------------------------------------------------
# Content sheet operations
# ---------------------------------------------------------------------------

def add_content(content_id: str, video_id: str, video_title: str,
                content_type: str, platform: str, content_text: str) -> None:
    """Append a generated content row with status 'Pending Approval'."""
    wb = get_workbook()
    ws = wb["Content"]
    ws.append([
        content_id, video_id, video_title, content_type,
        platform, content_text, "Pending Approval",
        "", "", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "",
    ])
    save_workbook(wb)


def get_approved_content() -> list[dict]:
    """Return all content rows with Status == 'Approved'."""
    wb = get_workbook()
    ws = wb["Content"]
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[6] == "Approved":
            result.append({
                "content_id":     row[0],
                "video_id":       row[1],
                "video_title":    row[2],
                "content_type":   row[3],
                "platform":       row[4],
                "content_text":   row[5],
                "scheduled_date": row[7],
            })
    return result


def update_content_status(content_id: str, new_status: str,
                          scheduled_date: str | None = None,
                          published_date: str | None = None) -> bool:
    """Update status (and optionally dates) for a content row."""
    wb = get_workbook()
    ws = wb["Content"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == content_id:
            row[6].value = new_status
            if scheduled_date:
                row[7].value = scheduled_date
            if published_date:
                row[8].value = published_date
            save_workbook(wb)
            return True
    return False


def list_content(status_filter: str | None = None) -> list[dict]:
    """Return content rows, optionally filtered by status."""
    wb = get_workbook()
    ws = wb["Content"]
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        if status_filter and row[6] != status_filter:
            continue
        result.append(dict(zip(CONTENT_HEADERS, row)))
    return result


# ---------------------------------------------------------------------------
# Config sheet operations
# ---------------------------------------------------------------------------

def get_config(key: str) -> str | None:
    """Read a config value by key."""
    wb = get_workbook()
    ws = wb["Config"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] == key:
            return str(row[1]) if row[1] is not None else None
    return None


def set_config(key: str, value: str) -> None:
    """Write a config value (update existing row or append new one)."""
    wb = get_workbook()
    ws = wb["Config"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == key:
            row[1].value = value
            save_workbook(wb)
            return
    ws.append([key, value, ""])
    save_workbook(wb)
